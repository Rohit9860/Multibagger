# """
# Screener.in Quarterly Results Scraper
# """

# import requests
# from bs4 import BeautifulSoup
# import pandas as pd
# from openpyxl import load_workbook
# from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# import time, re, logging, argparse, sys, random

# logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
# log = logging.getLogger(__name__)

# SCREENER_USERNAME = "mehetrerohit0@gmail.com"
# SCREENER_PASSWORD = "Rohit@9860"

# BASE_URL    = "https://www.screener.in"
# LOGIN_URL   = f"{BASE_URL}/login/"
# RESULTS_URL = f"{BASE_URL}/results/latest/"
# QUARTERS    = ["Mar 2025", "Dec 2025", "Mar 2026"]

# HEADERS = {
#     "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
#     "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
#     "Accept-Language": "en-IN,en;q=0.9",
#     "Accept-Encoding": "gzip, deflate, br",
#     "Connection": "keep-alive",
#     "Upgrade-Insecure-Requests": "1",
# }


# # ── Login ──────────────────────────────────────────────────────────────────────

# def login(username, password):
#     session = requests.Session()
#     session.headers.update(HEADERS)

#     log.info("Fetching login page for CSRF token...")
#     resp = session.get(LOGIN_URL, timeout=15)
#     if resp.status_code != 200:
#         log.error(f"Could not reach login page: HTTP {resp.status_code}")
#         sys.exit(1)

#     soup = BeautifulSoup(resp.text, "lxml")
#     csrf_input = soup.find("input", {"name": "csrfmiddlewaretoken"})
#     if not csrf_input:
#         csrf_meta = soup.find("meta", {"name": "csrf-token"})
#         csrf_token = csrf_meta["content"] if csrf_meta else session.cookies.get("csrftoken", "")
#     else:
#         csrf_token = csrf_input.get("value", "")

#     if not csrf_token:
#         csrf_token = session.cookies.get("csrftoken", "")

#     log.info(f"CSRF token: {csrf_token[:10]}..." if csrf_token else "CSRF token not found")

#     payload = {
#         "username":            username,
#         "password":            password,
#         "csrfmiddlewaretoken": csrf_token,
#         "next":                "/",
#     }
#     login_resp = session.post(
#         LOGIN_URL,
#         data=payload,
#         headers={**HEADERS, "Referer": LOGIN_URL, "Origin": BASE_URL},
#         timeout=15,
#         allow_redirects=True,
#     )

#     if "logout" in login_resp.text.lower() or "my account" in login_resp.text.lower():
#         log.info("✅ Login successful!")
#     elif "invalid" in login_resp.text.lower() or "incorrect" in login_resp.text.lower():
#         log.error("❌ Login failed: Invalid username or password.")
#         sys.exit(1)
#     elif login_resp.url and "login" not in login_resp.url:
#         log.info("✅ Login successful (redirected away from login page)")
#     else:
#         log.warning(f"⚠️  Login status unclear — continuing anyway... Final URL: {login_resp.url}")

#     return session


# # ── Fetch with exponential backoff ─────────────────────────────────────────────

# def fetch_page(session, page):
#     # NOTE: paginator uses ?p= not ?page=
#     url = RESULTS_URL if page == 1 else f"{RESULTS_URL}?p={page}"

#     for attempt in range(6):
#         try:
#             resp = session.get(url, timeout=20)

#             if resp.status_code == 429:
#                 retry_after = resp.headers.get("Retry-After")
#                 wait = int(retry_after) if retry_after else (2 ** attempt) * 5
#                 log.warning(f"429 on page {page} attempt {attempt+1} — waiting {wait}s")
#                 time.sleep(wait)
#                 continue

#             if resp.status_code == 403:
#                 log.error("403 Forbidden — session expired.")
#                 sys.exit(1)

#             resp.raise_for_status()
#             return resp.text

#         except requests.HTTPError as e:
#             wait = (2 ** attempt) * 3
#             log.warning(f"HTTP error page {page} attempt {attempt+1}: {e} — retrying in {wait}s")
#             time.sleep(wait)
#         except Exception as e:
#             wait = (2 ** attempt) * 3
#             log.warning(f"Request error page {page} attempt {attempt+1}: {e} — retrying in {wait}s")
#             time.sleep(wait)

#     log.error(f"❌ Failed to fetch page {page} after 6 attempts — skipping")
#     return ""


# # ── Utilities ──────────────────────────────────────────────────────────────────

# def parse_num(text):
#     if not text or str(text).strip() in ("-", "—", ""):
#         return None
#     cleaned = re.sub(r"[₹,\s%↑↓+⇡⇣]", "", str(text).strip())
#     try:
#         return float(cleaned)
#     except ValueError:
#         return None


# def get_total_pages(html):
#     """Parse total pages from already-fetched HTML. Paginator uses ?p=N."""
#     soup = BeautifulSoup(html, "lxml")
#     pages = set()
#     for a in soup.find_all("a", href=True):
#         m = re.search(r"[?&]p=(\d+)", a["href"])
#         if m:
#             pages.add(int(m.group(1)))
#     total = max(pages) if pages else 1
#     log.info(f"Total pages detected: {total}")
#     return total


# # ── HTML Parsing ───────────────────────────────────────────────────────────────

# def parse_companies_from_html(html):
#     """
#     Page structure per company:
#       <div class="flex-row ...">
#         <div>
#           <a href="/company/TICKER/...">Company Name</a>
#         </div>
#         ...price info...
#       </div>
#       <div class="bg-base border-radius-8 ...">
#         <table class="data-table">
#           <thead><tr><th></th><th>YOY</th><th>Mar 2026</th><th>Dec 2025</th><th>Mar 2025</th></tr></thead>
#           <tbody>
#             <tr><td>Sales</td><td>YOY%</td><td>val</td><td>val</td><td>val</td></tr>
#             <tr><td>EBIDT</td>...</tr>
#             <tr><td>Net profit</td>...</tr>
#             <tr><td>EPS</td>...</tr>
#           </tbody>
#         </table>
#       </div>
#     """
#     if not html:
#         return []

#     soup = BeautifulSoup(html, "lxml")
#     companies = []

#     # Find all result tables (class="data-table")
#     for table in soup.find_all("table", class_="data-table"):
#         thead = table.find("thead")
#         if not thead:
#             continue

#         header_cells = [th.get_text(strip=True) for th in thead.find_all("th")]
#         if "YOY" not in header_cells:
#             continue

#         # Map col index -> quarter name (col 0 = metric label, col 1 = YOY, col 2+ = quarters)
#         quarter_cols = {}
#         for i, h in enumerate(header_cells):
#             if i >= 2 and h:
#                 quarter_cols[i] = h

#         # ── Find company name ────────────────────────────────────────────────
#         # The table is inside: <div class="bg-base..."><table>
#         # Above that div is: <div class="flex-row..."><a href="/company/...">Name</a>
#         company_name = None
#         company_link = None

#         # Walk up to the wrapper div that contains the table
#         table_wrapper = table.find_parent("div", class_="bg-base")
#         if table_wrapper:
#             # The company heading div is the previous sibling of the wrapper
#             heading_div = table_wrapper.find_previous_sibling("div")
#             if heading_div:
#                 a = heading_div.find("a", href=lambda h: h and "/company/" in h)
#                 if a:
#                     company_name = a.find("span", class_="hover-link").get_text(strip=True) if a.find("span", class_="hover-link") else a.get_text(strip=True)
#                     href = a["href"]
#                     company_link = BASE_URL + href if href.startswith("/") else href

#         if not company_name:
#             log.debug(f"Could not find company name for a table — skipping")
#             continue

#         tbody = table.find("tbody")
#         if not tbody:
#             continue

#         # Find or create record for this company
#         record = next((c for c in companies if c["Company"] == company_name), None)
#         if record is None:
#             record = {"Company": company_name, "Link": company_link}
#             companies.append(record)

#         for row in tbody.find_all("tr"):
#             cells = row.find_all("td")
#             if not cells:
#                 continue
#             metric = cells[0].get_text(strip=True)
#             if not metric:
#                 continue
#             for col_idx, quarter in quarter_cols.items():
#                 if col_idx < len(cells):
#                     # Get raw text, strip any nested spans (e.g. ₹ symbol in EPS)
#                     val = cells[col_idx].get_text(strip=True)
#                     key = f"{metric}|{quarter}"
#                     record[key] = val

#     log.info(f"  Parsed {len(companies)} companies")
#     if companies:
#         log.info(f"  Sample: {companies[0]['Company']} | keys: {[k for k in companies[0] if k not in ('Company','Link')]}")
#     return companies


# # ── Filtering ──────────────────────────────────────────────────────────────────

# def extract_metric(record, keyword):
#     """Extract values for all QUARTERS for a given metric keyword."""
#     result = {}
#     for q in QUARTERS:
#         for k, v in record.items():
#             if keyword.lower() in k.lower() and q.lower() in k.lower():
#                 result[q] = parse_num(v)
#                 break
#     return result


# def is_strictly_increasing(vals):
#     v = [vals.get(q) for q in QUARTERS]
#     if any(x is None for x in v):
#         return False
#     return v[0] < v[1] < v[2]


# # ── Debug ──────────────────────────────────────────────────────────────────────

# def run_debug(session):
#     log.info("Debug mode — fetching page 1...")
#     html = fetch_page(session, 1)
#     with open("debug.html", "w", encoding="utf-8") as f:
#         f.write(html)
#     log.info("Saved → debug.html")

#     companies = parse_companies_from_html(html)
#     log.info(f"Parsed {len(companies)} companies from page 1")
#     for c in companies[:5]:
#         log.info(f"  {c.get('Company')} | Sales Mar 2025: {c.get('Sales|Mar 2025')} | Net profit|Mar 2025: {c.get('Net profit|Mar 2025')}")


# # ── Main Scraper ───────────────────────────────────────────────────────────────

# def run_scraper(session, max_pages=None, delay=3.0):
#     log.info("Fetching page 1...")
#     first_html = fetch_page(session, 1)
#     if not first_html:
#         log.error("Could not fetch page 1.")
#         return pd.DataFrame()

#     total = get_total_pages(first_html)
#     if max_pages:
#         total = min(total, max_pages)
#     log.info(f"Scraping {total} pages...")

#     all_companies = []

#     log.info(f"Page 1/{total}...")
#     companies = parse_companies_from_html(first_html)
#     log.info(f"  Got {len(companies)} companies")
#     all_companies.extend(companies)

#     for page in range(2, total + 1):
#         jitter = random.uniform(0.8, 1.5)
#         sleep_time = delay * jitter
#         log.info(f"Page {page}/{total} (waiting {sleep_time:.1f}s)...")
#         time.sleep(sleep_time)

#         html = fetch_page(session, page)
#         companies = parse_companies_from_html(html)
#         log.info(f"  Got {len(companies)} companies")
#         all_companies.extend(companies)

#     log.info(f"Total scraped: {len(all_companies)}")
#     if not all_companies:
#         log.error("0 companies found. Run with --debug to inspect HTML.")
#         return pd.DataFrame()

#     qualifying = []
#     for comp in all_companies:
#         sales  = extract_metric(comp, "Sales")
#         profit = extract_metric(comp, "Net profit")  # matches "Net profit" in HTML
#         if not is_strictly_increasing(sales) or not is_strictly_increasing(profit):
#             continue
#         s1, s2, s3 = sales["Mar 2025"], sales["Dec 2025"], sales["Mar 2026"]
#         p1, p2, p3 = profit["Mar 2025"], profit["Dec 2025"], profit["Mar 2026"]

#         # Skip if base quarter is zero (can't compute growth %)
#         if s1 == 0 or p1 == 0:
#             continue

#         qualifying.append({
#             "Company":             comp["Company"],
#             "Link":                comp["Link"],
#             "Sales Mar 2025":      s1,
#             "Sales Dec 2025":      s2,
#             "Sales Mar 2026":      s3,
#             "Net Profit Mar 2025": p1,
#             "Net Profit Dec 2025": p2,
#             "Net Profit Mar 2026": p3,
#             "Sales Growth %":      round((s3 - s1) / abs(s1) * 100, 2),
#             "Profit Growth %":     round((p3 - p1) / abs(p1) * 100, 2),
#         })

#     log.info(f"Qualifying companies: {len(qualifying)}")
#     return pd.DataFrame(qualifying)


# # ── Excel Export ───────────────────────────────────────────────────────────────

# def save_to_excel(df, output_path):
#     if df.empty:
#         log.warning("No qualifying companies found. Saving empty file.")

#     df.to_excel(output_path, index=False, sheet_name="Qualifying Companies")
#     wb = load_workbook(output_path)
#     ws = wb.active

#     hfill  = PatternFill("solid", fgColor="1F4E79")
#     hfont  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
#     gfill  = PatternFill("solid", fgColor="E2EFDA")
#     afill  = PatternFill("solid", fgColor="F0F7FF")
#     thin   = Side(style="thin", color="D0D0D0")
#     border = Border(left=thin, right=thin, top=thin, bottom=thin)

#     for cell in ws[1]:
#         cell.font      = hfont
#         cell.fill      = hfill
#         cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#         cell.border    = border
#     ws.row_dimensions[1].height = 36

#     for ri, row in enumerate(ws.iter_rows(min_row=2), start=2):
#         fill = gfill if ri % 2 == 0 else afill
#         for cell in row:
#             cell.fill      = fill
#             cell.alignment = Alignment(horizontal="left", vertical="center")
#             cell.border    = border

#     for col, w in {"A": 30, "B": 55, "C": 15, "D": 15, "E": 15,
#                    "F": 18, "G": 18, "H": 18, "I": 16, "J": 16}.items():
#         ws.column_dimensions[col].width = w

#     for ri in range(2, ws.max_row + 1):
#         cell = ws[f"B{ri}"]
#         if cell.value and str(cell.value).startswith("http"):
#             cell.hyperlink = cell.value
#             cell.font = Font(color="0563C1", underline="single", name="Arial")
#         for col in ["C","D","E","F","G","H"]:
#             ws[f"{col}{ri}"].number_format = "#,##0.0"
#         for col in ["I","J"]:
#             ws[f"{col}{ri}"].number_format = "0.00"

#     ws.freeze_panes = "A2"
#     wb.save(output_path)
#     log.info(f"✅ Saved → {output_path}")


# # ── Entry Point ────────────────────────────────────────────────────────────────

# if __name__ == "__main__":
#     parser = argparse.ArgumentParser()
#     parser.add_argument("--pages",    type=int,   default=None)
#     parser.add_argument("--delay",    type=float, default=3.0)
#     parser.add_argument("--output",   type=str,   default="screener_filtered_results.xlsx")
#     parser.add_argument("--username", type=str,   default=None)
#     parser.add_argument("--password", type=str,   default=None)
#     parser.add_argument("--debug",    action="store_true")
#     args = parser.parse_args()

#     username = args.username or SCREENER_USERNAME
#     password = args.password or SCREENER_PASSWORD

#     if username == "your_email@example.com":
#         log.error("Please set SCREENER_USERNAME and SCREENER_PASSWORD at the top of this file.")
#         sys.exit(1)

#     session = login(username, password)

#     if args.debug:
#         run_debug(session)
#         sys.exit(0)

#     df = run_scraper(session, max_pages=args.pages, delay=args.delay)
#     save_to_excel(df, args.output)
#     print(f"\n✅ Done! {len(df)} qualifying companies saved to {args.output}")



##################################################### SECOND ###########################################################



# """
# scraper.py — All scraping, parsing, filtering, and Excel-building logic.
# Import this module from app.py (or any other interface).
# """

# import io
# import logging
# import random
# import re
# import time

# import pandas as pd
# import requests
# from bs4 import BeautifulSoup
# from openpyxl import load_workbook, Workbook
# from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
# from openpyxl.utils import get_column_letter

# logging.basicConfig(level=logging.INFO)
# log = logging.getLogger(__name__)

# # ── Constants ──────────────────────────────────────────────────────────────────
# BASE_URL    = "https://www.screener.in"
# LOGIN_URL   = f"{BASE_URL}/login/"
# RESULTS_URL = f"{BASE_URL}/results/latest/"

# HEADERS = {
#     "User-Agent": (
#         "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
#         "AppleWebKit/537.36 Chrome/124.0.0.0 Safari/537.36"
#     ),
#     "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
#     "Accept-Language": "en-IN,en;q=0.9",
#     "Connection": "keep-alive",
# }


# # ── Auth ───────────────────────────────────────────────────────────────────────
# def login(username: str, password: str) -> requests.Session:
#     """Authenticate with Screener.in and return an authenticated session."""
#     session = requests.Session()
#     session.headers.update(HEADERS)

#     resp = session.get(LOGIN_URL, timeout=15)
#     if resp.status_code != 200:
#         raise RuntimeError(f"Cannot reach login page: HTTP {resp.status_code}")

#     soup = BeautifulSoup(resp.text, "lxml")
#     csrf = ""
#     inp = soup.find("input", {"name": "csrfmiddlewaretoken"})
#     if inp:
#         csrf = inp.get("value", "")
#     else:
#         meta = soup.find("meta", {"name": "csrf-token"})
#         csrf = meta["content"] if meta else session.cookies.get("csrftoken", "")
#     if not csrf:
#         csrf = session.cookies.get("csrftoken", "")

#     r = session.post(
#         LOGIN_URL,
#         data={
#             "username": username,
#             "password": password,
#             "csrfmiddlewaretoken": csrf,
#             "next": "/",
#         },
#         headers={**HEADERS, "Referer": LOGIN_URL, "Origin": BASE_URL},
#         timeout=15,
#         allow_redirects=True,
#     )
#     t = r.text.lower()
#     if "logout" in t or "my account" in t:
#         return session
#     if "invalid" in t or "incorrect" in t:
#         raise RuntimeError("Login failed — check your credentials.")
#     if r.url and "login" not in r.url:
#         return session
#     raise RuntimeError("Login status unclear — could not confirm authentication.")


# # ── HTTP helpers ───────────────────────────────────────────────────────────────
# def fetch(session: requests.Session, url: str) -> str:
#     """GET a URL with exponential backoff on errors / rate-limits."""
#     for attempt in range(6):
#         try:
#             resp = session.get(url, timeout=20)
#             if resp.status_code == 429:
#                 wait = int(resp.headers.get("Retry-After", (2 ** attempt) * 5))
#                 log.warning(f"429 on {url} — waiting {wait}s")
#                 time.sleep(wait)
#                 continue
#             if resp.status_code == 403:
#                 raise RuntimeError("403 Forbidden — session may have expired.")
#             resp.raise_for_status()
#             return resp.text
#         except RuntimeError:
#             raise
#         except Exception as e:
#             wait = (2 ** attempt) * 3
#             log.warning(f"Fetch error {url} attempt {attempt + 1}: {e} — retry in {wait}s")
#             time.sleep(wait)
#     return ""


# def fetch_page(session: requests.Session, page: int) -> str:
#     url = RESULTS_URL if page == 1 else f"{RESULTS_URL}?p={page}"
#     return fetch(session, url)


# # ── Parsing utilities ──────────────────────────────────────────────────────────
# def parse_num(text) -> float | None:
#     if not text or str(text).strip() in ("-", "—", ""):
#         return None
#     cleaned = re.sub(r"[₹,\s%↑↓⇡⇣+]", "", str(text).strip())
#     try:
#         return float(cleaned)
#     except ValueError:
#         return None


# def get_total_pages(html: str) -> int:
#     soup = BeautifulSoup(html, "lxml")
#     pages = set()
#     for a in soup.find_all("a", href=True):
#         m = re.search(r"[?&]p=(\d+)", a["href"])
#         if m:
#             pages.add(int(m.group(1)))
#     return max(pages) if pages else 1


# def detect_quarters_from_html(html: str) -> list:
#     """
#     Auto-detect the 3 most-recent quarter columns from the first data-table.
#     HTML order is newest→oldest; we reverse to chronological (oldest first).
#     """
#     soup = BeautifulSoup(html, "lxml")
#     for table in soup.find_all("table", class_="data-table"):
#         thead = table.find("thead")
#         if not thead:
#             continue
#         cells = [th.get_text(strip=True) for th in thead.find_all("th")]
#         quarters = [
#             c for c in cells
#             if re.match(r"(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{4}", c)
#         ]
#         if len(quarters) >= 3:
#             return list(reversed(quarters[:3]))   # oldest → newest
#     return []


# # ── Results page parser ────────────────────────────────────────────────────────
# def parse_companies_from_html(html: str, quarters: list) -> list:
#     """Parse all company cards from one results page."""
#     if not html or not quarters:
#         return []

#     soup = BeautifulSoup(html, "lxml")
#     companies = []

#     for table in soup.find_all("table", class_="data-table"):
#         thead = table.find("thead")
#         if not thead:
#             continue
#         header_cells = [th.get_text(strip=True) for th in thead.find_all("th")]
#         if "YOY" not in header_cells:
#             continue

#         quarter_cols = {i: h for i, h in enumerate(header_cells) if i >= 2 and h}

#         company_name = company_link = None
#         wrapper = table.find_parent("div", class_="bg-base")
#         if wrapper:
#             hdiv = wrapper.find_previous_sibling("div")
#             if hdiv:
#                 a = hdiv.find("a", href=lambda h: h and "/company/" in h)
#                 if a:
#                     span = a.find("span", class_="hover-link")
#                     company_name = span.get_text(strip=True) if span else a.get_text(strip=True)
#                     href = a["href"]
#                     company_link = BASE_URL + href if href.startswith("/") else href

#         if not company_name:
#             continue

#         tbody = table.find("tbody")
#         if not tbody:
#             continue

#         record = next((c for c in companies if c["Company"] == company_name), None)
#         if record is None:
#             record = {"Company": company_name, "Link": company_link}
#             companies.append(record)

#         for row in tbody.find_all("tr"):
#             cells = row.find_all("td")
#             if not cells:
#                 continue
#             metric = cells[0].get_text(strip=True)
#             if not metric:
#                 continue
#             for col_idx, quarter in quarter_cols.items():
#                 if col_idx < len(cells):
#                     record[f"{metric}|{quarter}"] = cells[col_idx].get_text(strip=True)

#     return companies


# def extract_metric(record: dict, keyword: str, quarters: list) -> dict:
#     result = {}
#     for q in quarters:
#         for k, v in record.items():
#             if keyword.lower() in k.lower() and q.lower() in k.lower():
#                 result[q] = parse_num(v)
#                 break
#     return result


# def is_strictly_increasing(vals: dict, quarters: list) -> bool:
#     v = [vals.get(q) for q in quarters]
#     if any(x is None for x in v):
#         return False
#     return v[0] < v[1] < v[2]


# # ── Deep company detail scraper ────────────────────────────────────────────────
# def scrape_company_detail(session: requests.Session, company_name: str, link: str) -> dict:
#     """
#     Fetch a company's Screener page and extract:
#     - Key ratios: ROE, ROCE, P/BV, PE, OPM, Interest Cover, Promoter Holding, Market Cap
#     - Quarterly P&L (last 6 quarters)
#     - Annual P&L (last 3 years)
#     """
#     html = fetch(session, link)
#     if not html:
#         return {"Company": company_name, "Link": link, "Error": "Failed to fetch"}

#     soup = BeautifulSoup(html, "lxml")
#     result = {"Company": company_name, "Link": link}

#     # ── Key ratios ─────────────────────────────────────────────────────────────
#     ratio_map = {
#         "Market Cap":       ["market cap", "mcap"],
#         "P/E":              ["stock p/e", "p/e"],
#         "P/BV":             ["price to book", "p/b"],
#         "ROE":              ["return on equity", "roe"],
#         "ROCE":             ["roce", "return on capital"],
#         "OPM":              ["opm", "operating profit margin"],
#         "Promoter Holding": ["promoter holding", "promoter"],
#         "Debt to Equity":   ["debt to equity", "d/e"],
#         "Interest Cover":   ["interest coverage", "interest cover"],
#     }
#     for section_id in ["top-ratios", "company-info"]:
#         section = soup.find(id=section_id)
#         if not section:
#             continue
#         for li in section.find_all("li"):
#             label_el = li.find("span", class_="name")
#             value_el = li.find("span", class_="nowrap") or li.find("span", class_="number")
#             if not label_el or not value_el:
#                 continue
#             label_text = label_el.get_text(strip=True).lower()
#             value_text = value_el.get_text(strip=True)
#             for col_name, keywords in ratio_map.items():
#                 if col_name not in result:
#                     if any(kw in label_text for kw in keywords):
#                         result[col_name] = parse_num(value_text)

#     # ── Quarterly P&L ──────────────────────────────────────────────────────────
#     qsec = soup.find(id="quarters")
#     if qsec:
#         qtable = qsec.find("table")
#         if qtable:
#             thead = qtable.find("thead")
#             tbody = qtable.find("tbody")
#             if thead and tbody:
#                 q_headers = [th.get_text(strip=True) for th in thead.find_all("th")]
#                 q_dates   = q_headers[1:]
#                 last_n    = min(6, len(q_dates))
#                 q_use     = q_dates[-last_n:]
#                 for row in tbody.find_all("tr"):
#                     cells  = row.find_all("td")
#                     if not cells:
#                         continue
#                     metric = cells[0].get_text(strip=True)
#                     if metric.lower() in ("sales", "net profit", "ebidt", "eps", "opm %"):
#                         offset = len(q_dates) - last_n + 1
#                         for i, qd in enumerate(q_use):
#                             ci = offset + i
#                             if ci < len(cells):
#                                 result[f"{metric} | {qd}"] = parse_num(
#                                     cells[ci].get_text(strip=True)
#                                 )

#     # ── Annual P&L ────────────────────────────────────────────────────────────
#     asec = soup.find(id="profit-loss")
#     if asec:
#         atable = asec.find("table")
#         if atable:
#             thead = atable.find("thead")
#             tbody = atable.find("tbody")
#             if thead and tbody:
#                 yr_headers = [th.get_text(strip=True) for th in thead.find_all("th")]
#                 yr_dates   = yr_headers[1:]
#                 last_n     = min(3, len(yr_dates))
#                 yr_use     = yr_dates[-last_n:]
#                 for row in tbody.find_all("tr"):
#                     cells  = row.find_all("td")
#                     if not cells:
#                         continue
#                     metric = cells[0].get_text(strip=True)
#                     if metric.lower() in ("sales", "net profit", "opm %"):
#                         offset = len(yr_dates) - last_n + 1
#                         for i, yr in enumerate(yr_use):
#                             ci = offset + i
#                             if ci < len(cells):
#                                 result[f"Annual {metric} {yr}"] = parse_num(
#                                     cells[ci].get_text(strip=True)
#                                 )

#     return result


# # ── Excel style helpers ────────────────────────────────────────────────────────
# def _hdr_style(ws, row: int = 1) -> None:
#     hfill = PatternFill("solid", fgColor="1F4E79")
#     hfont = Font(bold=True, color="FFFFFF", name="Arial", size=10)
#     thin  = Side(style="thin", color="B0B0B0")
#     bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
#     for cell in ws[row]:
#         cell.font      = hfont
#         cell.fill      = hfill
#         cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#         cell.border    = bdr
#     ws.row_dimensions[row].height = 32


# def _row_style(ws, start: int = 2) -> None:
#     gfill = PatternFill("solid", fgColor="E2EFDA")
#     afill = PatternFill("solid", fgColor="F0F7FF")
#     thin  = Side(style="thin", color="D0D0D0")
#     bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
#     for ri, row in enumerate(ws.iter_rows(min_row=start), start=start):
#         fill = gfill if ri % 2 == 0 else afill
#         for cell in row:
#             cell.fill      = fill
#             cell.alignment = Alignment(horizontal="left", vertical="center")
#             cell.border    = bdr


# def _autowidth(ws, max_w: int = 42) -> None:
#     for col_cells in ws.columns:
#         w = max((len(str(c.value or "")) for c in col_cells), default=10)
#         ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(w + 4, max_w)


# def _hyperlink_col(ws, col_name: str = "Link") -> None:
#     for c in ws[1]:
#         if str(c.value) == col_name:
#             col = c.column
#             for ri in range(2, ws.max_row + 1):
#                 cell = ws.cell(row=ri, column=col)
#                 if cell.value and str(cell.value).startswith("http"):
#                     cell.hyperlink = cell.value
#                     cell.font = Font(color="0563C1", underline="single", name="Arial", size=10)
#             return


# # ── Excel builders ─────────────────────────────────────────────────────────────
# def build_filtered_excel(df: pd.DataFrame, quarters: list) -> bytes:
#     """Build Excel 1 — filtered summary with growth columns."""
#     buf = io.BytesIO()
#     df.to_excel(buf, index=False, sheet_name="Filtered Companies")
#     buf.seek(0)
#     wb = load_workbook(buf)
#     ws = wb.active
#     _hdr_style(ws)
#     _row_style(ws)
#     _autowidth(ws)
#     _hyperlink_col(ws, "Link")

#     # Color growth columns
#     for col_cells in ws.columns:
#         hdr = str(ws.cell(row=1, column=col_cells[0].column).value or "")
#         if "Growth %" in hdr:
#             for ri in range(2, ws.max_row + 1):
#                 cell = ws.cell(row=ri, column=col_cells[0].column)
#                 try:
#                     v = float(cell.value)
#                     if v >= 20:
#                         fgColor, font_color, bold = "C6EFCE", "276221", True
#                     elif v >= 0:
#                         fgColor, font_color, bold = "FFEB9C", "9C5700", False
#                     else:
#                         fgColor, font_color, bold = "FFC7CE", "9C0006", False
#                     cell.fill          = PatternFill("solid", fgColor=fgColor)
#                     cell.font          = Font(color=font_color, bold=bold, name="Arial", size=10)
#                     cell.number_format = "0.00"
#                 except Exception:
#                     pass

#     ws.freeze_panes = "A2"

#     # Summary sheet
#     ws2 = wb.create_sheet("Summary")
#     ws2["A1"] = "Metric"
#     ws2["B1"] = "Value"
#     _hdr_style(ws2)
#     rows = [
#         ("Total Qualifying Companies", len(df)),
#         ("Quarters Used",              " → ".join(quarters)),
#         ("Avg Sales Growth %",
#          round(df["Sales Growth %"].mean(), 2) if "Sales Growth %" in df.columns else "—"),
#         ("Avg Profit Growth %",
#          round(df["Profit Growth %"].mean(), 2) if "Profit Growth %" in df.columns else "—"),
#         ("Max Sales Growth %",
#          round(df["Sales Growth %"].max(), 2)  if "Sales Growth %" in df.columns else "—"),
#         ("Max Profit Growth %",
#          round(df["Profit Growth %"].max(), 2) if "Profit Growth %" in df.columns else "—"),
#         ("Top Company (Profit Growth)",
#          df.nlargest(1, "Profit Growth %")["Company"].values[0] if len(df) else "—"),
#     ]
#     for i, (k, v) in enumerate(rows, start=2):
#         ws2[f"A{i}"] = k
#         ws2[f"B{i}"] = v
#     ws2.column_dimensions["A"].width = 34
#     ws2.column_dimensions["B"].width = 22

#     out = io.BytesIO()
#     wb.save(out)
#     return out.getvalue()


# def build_detailed_excel(records: list) -> bytes:
#     """Build Excel 2 — deep company analysis across three sheets."""
#     if not records:
#         wb = Workbook()
#         ws = wb.active
#         ws["A1"] = "No data"
#         out = io.BytesIO()
#         wb.save(out)
#         return out.getvalue()

#     df = pd.DataFrame(records)

#     ratio_cols = [
#         "Company", "Link", "Market Cap", "P/E", "P/BV", "ROE", "ROCE",
#         "OPM", "Interest Cover", "Debt to Equity", "Promoter Holding", "Error",
#     ]
#     ratio_cols = [c for c in ratio_cols if c in df.columns]
#     df_ratios  = df[ratio_cols].copy()

#     qtr_pat  = re.compile(r".+\|\s*(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{4}")
#     qtr_cols = ["Company"] + [c for c in df.columns if qtr_pat.match(str(c))]
#     df_qtr   = df[[c for c in qtr_cols if c in df.columns]].copy()

#     ann_cols = ["Company"] + [c for c in df.columns if str(c).startswith("Annual ")]
#     df_ann   = df[[c for c in ann_cols if c in df.columns]].copy()

#     buf = io.BytesIO()
#     with pd.ExcelWriter(buf, engine="openpyxl") as writer:
#         df_ratios.to_excel(writer, index=False, sheet_name="Key Ratios")
#         df_qtr.to_excel(writer,    index=False, sheet_name="Quarterly P&L")
#         df_ann.to_excel(writer,    index=False, sheet_name="Annual P&L")
#     buf.seek(0)

#     wb = load_workbook(buf)
#     for sname in wb.sheetnames:
#         ws = wb[sname]
#         _hdr_style(ws)
#         _row_style(ws)
#         _autowidth(ws)
#         if sname == "Key Ratios":
#             _hyperlink_col(ws, "Link")

#         thresholds = {"ROE": 15, "ROCE": 15, "OPM": 10}
#         for col_cells in ws.columns:
#             hdr = str(ws.cell(row=1, column=col_cells[0].column).value or "")
#             if hdr in thresholds:
#                 thr = thresholds[hdr]
#                 for ri in range(2, ws.max_row + 1):
#                     cell = ws.cell(row=ri, column=col_cells[0].column)
#                     try:
#                         v    = float(cell.value)
#                         good = v >= thr
#                         cell.fill          = PatternFill("solid", fgColor="C6EFCE" if good else "FFC7CE")
#                         cell.font          = Font(
#                             color="276221" if good else "9C0006",
#                             bold=True, name="Arial", size=10,
#                         )
#                         cell.number_format = "0.0"
#                     except Exception:
#                         pass
#         ws.freeze_panes = "A2"

#     out = io.BytesIO()
#     wb.save(out)
#     return out.getvalue()


# # ── Core orchestrator ──────────────────────────────────────────────────────────
# def run_full_scrape(
#     session: requests.Session,
#     max_pages: int,
#     delay: float,
#     on_page_start=None,   # optional callback(page, total)
#     on_page_done=None,    # optional callback(page, total, n_companies)
# ) -> tuple[pd.DataFrame, list]:
#     """
#     Scrape all results pages, filter for strictly-increasing Sales & Net Profit,
#     and return (DataFrame, quarters).

#     Callbacks receive progress info so callers (e.g. Streamlit) can update UI.
#     """
#     first_html = fetch_page(session, 1)
#     if not first_html:
#         raise RuntimeError("Could not fetch page 1.")

#     quarters = detect_quarters_from_html(first_html)
#     if not quarters:
#         raise RuntimeError(
#             "Could not detect quarters from HTML — check site structure."
#         )

#     total = get_total_pages(first_html)
#     if max_pages and max_pages > 0:
#         total = min(total, max_pages)

#     all_companies = []
#     if on_page_done:
#         on_page_done(1, total, len(all_companies))

#     all_companies.extend(parse_companies_from_html(first_html, quarters))

#     for page in range(2, total + 1):
#         sleep_time = delay * random.uniform(0.8, 1.5)
#         if on_page_start:
#             on_page_start(page, total, len(all_companies))
#         time.sleep(sleep_time)
#         html = fetch_page(session, page)
#         all_companies.extend(parse_companies_from_html(html, quarters))
#         if on_page_done:
#             on_page_done(page, total, len(all_companies))

#     qualifying = []
#     for comp in all_companies:
#         sales  = extract_metric(comp, "Sales",      quarters)
#         profit = extract_metric(comp, "Net profit", quarters)
#         if not is_strictly_increasing(sales,  quarters):
#             continue
#         if not is_strictly_increasing(profit, quarters):
#             continue
#         s = [sales[q]  for q in quarters]
#         p = [profit[q] for q in quarters]
#         if not s[0] or not p[0]:
#             continue
#         row = {"Company": comp["Company"], "Link": comp["Link"]}
#         for i, q in enumerate(quarters):
#             row[f"Sales {q}"]      = s[i]
#             row[f"Net Profit {q}"] = p[i]
#         row["Sales Growth %"]  = round((s[2] - s[0]) / abs(s[0]) * 100, 2)
#         row["Profit Growth %"] = round((p[2] - p[0]) / abs(p[0]) * 100, 2)
#         qualifying.append(row)

#     return pd.DataFrame(qualifying), quarters


##################################################### THIRD ###########################################################

"""
scraper.py — All scraping, parsing, filtering, and Excel-building logic.
Import this module from app.py (or any other interface).

Detail columns extracted per company
──────────────────────────────────────
ROE              → top-ratios (screener key ratios)
ROCE             → top-ratios (screener key ratios)
P/BV             → Current Price / Book Value  (both from top-ratios)
PE/EPS           → Stock P/E / latest quarterly EPS (Rs)
PAT Jumped CR    → Profit Growth % (passed in from filtered scrape)
OPM (Quarterly)  → latest OPM % from Quarterly P&L table
OPM (P&L)        → latest OPM % from Annual Profit & Loss table
Interest         → (latest Borrowings / latest Reserves) * 100  [Balance Sheet]
Promoter Latest  → latest promoter holding %  [Shareholding Pattern]
Promoter Prev    → second-latest promoter holding %
Promoter Δ       → Latest − Prev  (positive = promoters buying)
"""

import io
import logging
import random
import re
import time

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO)
log = logging.getLogger(__name__)

# ── Constants ──────────────────────────────────────────────────────────────────
BASE_URL    = "https://www.screener.in"
LOGIN_URL   = f"{BASE_URL}/login/"
RESULTS_URL = f"{BASE_URL}/results/latest/"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-IN,en;q=0.9",
    "Connection": "keep-alive",
}


# ── Auth ───────────────────────────────────────────────────────────────────────
def login(username: str, password: str) -> requests.Session:
    """Authenticate with Screener.in and return an authenticated session."""
    session = requests.Session()
    session.headers.update(HEADERS)

    resp = session.get(LOGIN_URL, timeout=15)
    if resp.status_code != 200:
        raise RuntimeError(f"Cannot reach login page: HTTP {resp.status_code}")

    soup = BeautifulSoup(resp.text, "lxml")
    csrf = ""
    inp = soup.find("input", {"name": "csrfmiddlewaretoken"})
    if inp:
        csrf = inp.get("value", "")
    else:
        meta = soup.find("meta", {"name": "csrf-token"})
        csrf = meta["content"] if meta else session.cookies.get("csrftoken", "")
    if not csrf:
        csrf = session.cookies.get("csrftoken", "")

    r = session.post(
        LOGIN_URL,
        data={
            "username": username,
            "password": password,
            "csrfmiddlewaretoken": csrf,
            "next": "/",
        },
        headers={**HEADERS, "Referer": LOGIN_URL, "Origin": BASE_URL},
        timeout=15,
        allow_redirects=True,
    )
    t = r.text.lower()
    if "logout" in t or "my account" in t:
        return session
    if "invalid" in t or "incorrect" in t:
        raise RuntimeError("Login failed — check your credentials.")
    if r.url and "login" not in r.url:
        return session
    raise RuntimeError("Login status unclear — could not confirm authentication.")


# ── HTTP helpers ───────────────────────────────────────────────────────────────
def fetch(session: requests.Session, url: str) -> str:
    """GET a URL with exponential backoff on errors / rate-limits."""
    for attempt in range(6):
        try:
            resp = session.get(url, timeout=20)
            if resp.status_code == 429:
                wait = int(resp.headers.get("Retry-After", (2 ** attempt) * 5))
                log.warning(f"429 on {url} — waiting {wait}s")
                time.sleep(wait)
                continue
            if resp.status_code == 403:
                raise RuntimeError("403 Forbidden — session may have expired.")
            resp.raise_for_status()
            return resp.text
        except RuntimeError:
            raise
        except Exception as e:
            wait = (2 ** attempt) * 3
            log.warning(f"Fetch error {url} attempt {attempt + 1}: {e} — retry in {wait}s")
            time.sleep(wait)
    return ""


def fetch_page(session: requests.Session, page: int) -> str:
    url = RESULTS_URL if page == 1 else f"{RESULTS_URL}?p={page}"
    return fetch(session, url)


# ── Parsing utilities ──────────────────────────────────────────────────────────
def parse_num(text) -> float | None:
    if not text or str(text).strip() in ("-", "—", ""):
        return None
    cleaned = re.sub(r"[₹,\s%↑↓⇡⇣+]", "", str(text).strip())
    try:
        return float(cleaned)
    except ValueError:
        return None


def get_total_pages(html: str) -> int:
    soup = BeautifulSoup(html, "lxml")
    pages = set()
    for a in soup.find_all("a", href=True):
        m = re.search(r"[?&]p=(\d+)", a["href"])
        if m:
            pages.add(int(m.group(1)))
    return max(pages) if pages else 1


def detect_quarters_from_html(html: str) -> list:
    """
    Auto-detect the 3 most-recent quarter columns from the first data-table.
    HTML order is newest→oldest; we reverse to chronological (oldest first).
    """
    soup = BeautifulSoup(html, "lxml")
    for table in soup.find_all("table", class_="data-table"):
        thead = table.find("thead")
        if not thead:
            continue
        cells = [th.get_text(strip=True) for th in thead.find_all("th")]
        quarters = [
            c for c in cells
            if re.match(r"(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{4}", c)
        ]
        if len(quarters) >= 3:
            return list(reversed(quarters[:3]))   # oldest → newest
    return []


# ── Results page parser ────────────────────────────────────────────────────────
def parse_companies_from_html(html: str, quarters: list) -> list:
    """Parse all company cards from one results page."""
    if not html or not quarters:
        return []

    soup = BeautifulSoup(html, "lxml")
    companies = []

    for table in soup.find_all("table", class_="data-table"):
        thead = table.find("thead")
        if not thead:
            continue
        header_cells = [th.get_text(strip=True) for th in thead.find_all("th")]
        if "YOY" not in header_cells:
            continue

        quarter_cols = {i: h for i, h in enumerate(header_cells) if i >= 2 and h}

        company_name = company_link = None
        wrapper = table.find_parent("div", class_="bg-base")
        if wrapper:
            hdiv = wrapper.find_previous_sibling("div")
            if hdiv:
                a = hdiv.find("a", href=lambda h: h and "/company/" in h)
                if a:
                    span = a.find("span", class_="hover-link")
                    company_name = span.get_text(strip=True) if span else a.get_text(strip=True)
                    href = a["href"]
                    company_link = BASE_URL + href if href.startswith("/") else href

        if not company_name:
            continue

        tbody = table.find("tbody")
        if not tbody:
            continue

        record = next((c for c in companies if c["Company"] == company_name), None)
        if record is None:
            record = {"Company": company_name, "Link": company_link}
            companies.append(record)

        for row in tbody.find_all("tr"):
            cells = row.find_all("td")
            if not cells:
                continue
            metric = cells[0].get_text(strip=True)
            if not metric:
                continue
            for col_idx, quarter in quarter_cols.items():
                if col_idx < len(cells):
                    record[f"{metric}|{quarter}"] = cells[col_idx].get_text(strip=True)

    return companies


def extract_metric(record: dict, keyword: str, quarters: list) -> dict:
    result = {}
    for q in quarters:
        for k, v in record.items():
            if keyword.lower() in k.lower() and q.lower() in k.lower():
                result[q] = parse_num(v)
                break
    return result


def is_strictly_increasing(vals: dict, quarters: list) -> bool:
    v = [vals.get(q) for q in quarters]
    if any(x is None for x in v):
        return False
    return v[0] < v[1] < v[2]


# ── HTML table helpers ─────────────────────────────────────────────────────────
def _table_to_row_dict(table_tag) -> dict[str, list]:
    """
    Convert a screener data-table into {row_label_lower: [val1, val2, ...]}
    Values are left→right as they appear in HTML (screener: newest = rightmost).
    """
    if not table_tag:
        return {}
    result: dict[str, list] = {}
    tbody = table_tag.find("tbody")
    if not tbody:
        return {}
    for tr in tbody.find_all("tr"):
        cells = tr.find_all("td")
        if not cells:
            continue
        label = cells[0].get_text(strip=True).lower().strip()
        vals  = [cells[i].get_text(strip=True) for i in range(1, len(cells))]
        if label:
            result[label] = vals
    return result


def _table_headers(table_tag) -> list:
    """Return thead column labels (excluding the first label column)."""
    if not table_tag:
        return []
    thead = table_tag.find("thead")
    if not thead:
        return []
    ths = thead.find_all("th")
    return [th.get_text(strip=True) for th in ths[1:]]


# ── Deep company detail scraper ────────────────────────────────────────────────
def scrape_company_detail(
    session: requests.Session,
    company_name: str,
    link: str,
    pat_growth: float | None = None,
) -> dict:
    """
    Scrape a single company page and return a dict with these keys:
      Company, Link,
      ROE, ROCE, P/BV, PE/EPS, PAT Jumped CR,
      OPM (Quarterly), OPM (P&L),
      Interest,
      Promoter Latest %, Promoter Prev %, Promoter Δ,
      Error (only if something failed)
    """
    html = fetch(session, link)
    if not html:
        return {"Company": company_name, "Link": link, "Error": "Failed to fetch"}

    soup = BeautifulSoup(html, "lxml")
    result: dict = {"Company": company_name, "Link": link}

    # ── 1. Key ratios (top-ratios / company-info section) ─────────────────────
    ratio_raw: dict[str, str] = {}
    for section_id in ["top-ratios", "company-info"]:
        section = soup.find(id=section_id)
        if not section:
            continue
        for li in section.find_all("li"):
            label_el = li.find("span", class_="name")
            value_el = (
                li.find("span", class_="nowrap")
                or li.find("span", class_="number")
            )
            if not label_el or not value_el:
                continue
            lbl = label_el.get_text(strip=True).lower()
            val = value_el.get_text(strip=True)
            ratio_raw[lbl] = val

    def _get_ratio(*keywords) -> float | None:
        for kw in keywords:
            for lbl, val in ratio_raw.items():
                if kw in lbl:
                    return parse_num(val)
        return None

    roe           = _get_ratio("return on equity", "roe")
    roce          = _get_ratio("roce", "return on capital")
    current_price = _get_ratio("current price")
    book_value    = _get_ratio("book value")
    stock_pe      = _get_ratio("stock p/e", "p/e")

    result["ROE"]  = roe
    result["ROCE"] = roce

    # P/BV = Current Price / Book Value
    if current_price and book_value and book_value != 0:
        result["P/BV"] = round(current_price / book_value, 4)
    else:
        result["P/BV"] = None

    # PAT Jumped CR = Profit Growth % passed in from filtered df
    result["PAT Jumped CR"] = pat_growth

    # ── 2. Quarterly P&L — latest EPS (Rs) and latest OPM % ──────────────────
    latest_eps_q = None
    latest_opm_q = None

    qsec = soup.find(id="quarters")
    if qsec:
        qtable = qsec.find("table")
        if qtable:
            rows_q = _table_to_row_dict(qtable)
            # Screener quarterly table: columns left→right = oldest→newest
            # So vals[-1] = most recent quarter
            for label, vals in rows_q.items():
                if not vals:
                    continue
                latest_val = vals[-1]
                # EPS row is usually "eps in rs" or "eps"
                if "eps" in label:
                    latest_eps_q = parse_num(latest_val)
                elif label in ("opm %", "opm%"):
                    latest_opm_q = parse_num(latest_val)

    result["OPM (Quarterly)"] = latest_opm_q

    # PE/EPS = Stock P/E  ÷  latest quarterly EPS (Rs)
    if stock_pe is not None and latest_eps_q and latest_eps_q != 0:
        result["PE/EPS"] = round(stock_pe / latest_eps_q, 4)
    else:
        result["PE/EPS"] = stock_pe   # fallback: just P/E if EPS unavailable

    # ── 3. Annual Profit & Loss — latest OPM % ───────────────────────────────
    latest_opm_pl = None

    pl_sec = soup.find(id="profit-loss")
    if pl_sec:
        pl_table = pl_sec.find("table")
        if pl_table:
            rows_pl = _table_to_row_dict(pl_table)
            for label, vals in rows_pl.items():
                if label in ("opm %", "opm%") and vals:
                    latest_opm_pl = parse_num(vals[-1])
                    break

    result["OPM (P&L)"] = latest_opm_pl

    # ── 4. Balance Sheet — Interest = (Borrowings / Reserves) × 100 ──────────
    borrowings = None
    reserves   = None

    bs_sec = soup.find(id="balance-sheet")
    if bs_sec:
        bs_table = bs_sec.find("table")
        if bs_table:
            rows_bs = _table_to_row_dict(bs_table)
            for label, vals in rows_bs.items():
                if not vals:
                    continue
                if "borrowings" in label:
                    borrowings = parse_num(vals[-1])
                elif "reserves" in label:
                    reserves = parse_num(vals[-1])

    if borrowings is not None and reserves and reserves != 0:
        result["Interest"] = round((borrowings / reserves) * 100, 2)
    else:
        result["Interest"] = None

    # ── 5. Shareholding Pattern — latest 2 promoter values ───────────────────
    promoter_latest = None
    promoter_prev   = None

    # Try common section ids
    sh_sec = soup.find(id="shareholding")
    if not sh_sec:
        for tag in soup.find_all(id=re.compile(r"shareholding", re.I)):
            sh_sec = tag
            break

    if sh_sec:
        sh_table = sh_sec.find("table")
        if sh_table:
            rows_sh = _table_to_row_dict(sh_table)
            for label, vals in rows_sh.items():
                if "promoter" in label:
                    clean_vals = [parse_num(v) for v in vals if parse_num(v) is not None]
                    if len(clean_vals) >= 2:
                        promoter_latest = clean_vals[-1]
                        promoter_prev   = clean_vals[-2]
                    elif len(clean_vals) == 1:
                        promoter_latest = clean_vals[0]
                    break  # first promoter row is sufficient

    result["Promoter Latest %"] = promoter_latest
    result["Promoter Prev %"]   = promoter_prev
    result["Promoter Δ"] = (
        round(promoter_latest - promoter_prev, 2)
        if (promoter_latest is not None and promoter_prev is not None)
        else None
    )

    return result


# ── Excel style helpers ────────────────────────────────────────────────────────
def _hdr_style(ws, row: int = 1) -> None:
    hfill = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    thin  = Side(style="thin", color="B0B0B0")
    bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[row]:
        cell.font      = hfont
        cell.fill      = hfill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = bdr
    ws.row_dimensions[row].height = 32


def _row_style(ws, start: int = 2) -> None:
    gfill = PatternFill("solid", fgColor="E2EFDA")
    afill = PatternFill("solid", fgColor="F0F7FF")
    thin  = Side(style="thin", color="D0D0D0")
    bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ri, row in enumerate(ws.iter_rows(min_row=start), start=start):
        fill = gfill if ri % 2 == 0 else afill
        for cell in row:
            cell.fill      = fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border    = bdr


def _autowidth(ws, max_w: int = 42) -> None:
    for col_cells in ws.columns:
        w = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(w + 4, max_w)


def _hyperlink_col(ws, col_name: str = "Link") -> None:
    for c in ws[1]:
        if str(c.value) == col_name:
            col = c.column
            for ri in range(2, ws.max_row + 1):
                cell = ws.cell(row=ri, column=col)
                if cell.value and str(cell.value).startswith("http"):
                    cell.hyperlink = cell.value
                    cell.font = Font(color="0563C1", underline="single", name="Arial", size=10)
            return


def _color_cell(cell, value: float | None, low: float, high: float,
                reverse: bool = False) -> None:
    """
    Colour a cell green / yellow / red.
    reverse=True  →  lower values are better (e.g. P/BV, Interest).
    """
    if value is None:
        return
    try:
        v = float(value)
        if not reverse:
            if v >= high:
                fg, fc, bold = "C6EFCE", "276221", True
            elif v >= low:
                fg, fc, bold = "FFEB9C", "9C5700", False
            else:
                fg, fc, bold = "FFC7CE", "9C0006", False
        else:
            if v <= low:
                fg, fc, bold = "C6EFCE", "276221", True
            elif v <= high:
                fg, fc, bold = "FFEB9C", "9C5700", False
            else:
                fg, fc, bold = "FFC7CE", "9C0006", False
        cell.fill = PatternFill("solid", fgColor=fg)
        cell.font = Font(color=fc, bold=bold, name="Arial", size=10)
    except Exception:
        pass


# ── Excel builders ─────────────────────────────────────────────────────────────
def build_filtered_excel(df: pd.DataFrame, quarters: list) -> bytes:
    """Build Excel 1 — filtered summary with growth columns."""
    buf = io.BytesIO()
    df.to_excel(buf, index=False, sheet_name="Filtered Companies")
    buf.seek(0)
    wb = load_workbook(buf)
    ws = wb.active
    _hdr_style(ws)
    _row_style(ws)
    _autowidth(ws)
    _hyperlink_col(ws, "Link")

    for col_cells in ws.columns:
        hdr = str(ws.cell(row=1, column=col_cells[0].column).value or "")
        if "Growth %" in hdr:
            for ri in range(2, ws.max_row + 1):
                cell = ws.cell(row=ri, column=col_cells[0].column)
                v = parse_num(str(cell.value or ""))
                _color_cell(cell, v, low=0, high=20)
                if v is not None:
                    cell.number_format = "0.00"

    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Metric"
    ws2["B1"] = "Value"
    _hdr_style(ws2)
    summary_rows = [
        ("Total Qualifying Companies", len(df)),
        ("Quarters Used",              " → ".join(quarters)),
        ("Avg Sales Growth %",
         round(df["Sales Growth %"].mean(), 2) if "Sales Growth %" in df.columns else "—"),
        ("Avg Profit Growth %",
         round(df["Profit Growth %"].mean(), 2) if "Profit Growth %" in df.columns else "—"),
        ("Max Sales Growth %",
         round(df["Sales Growth %"].max(), 2)   if "Sales Growth %" in df.columns else "—"),
        ("Max Profit Growth %",
         round(df["Profit Growth %"].max(), 2)  if "Profit Growth %" in df.columns else "—"),
        ("Top Company (Profit Growth)",
         df.nlargest(1, "Profit Growth %")["Company"].values[0] if len(df) else "—"),
    ]
    for i, (k, v) in enumerate(summary_rows, start=2):
        ws2[f"A{i}"] = k
        ws2[f"B{i}"] = v
    ws2.column_dimensions["A"].width = 34
    ws2.column_dimensions["B"].width = 22

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def build_detailed_excel(records: list) -> bytes:
    """
    Build Excel 2 — single sheet with these exact columns:
    Company | Link | ROE | ROCE | P/BV | PE/EPS | PAT Jumped CR
            | OPM (Quarterly) | OPM (P&L)
            | Interest
            | Promoter Latest % | Promoter Prev % | Promoter Δ
            | Error (if any)
    """
    if not records:
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "No data"
        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

    df = pd.DataFrame(records)

    ordered = [
        "Company", "Link",
        "ROE", "ROCE", "P/BV",
        "PE/EPS",
        "PAT Jumped CR",
        "OPM (Quarterly)", "OPM (P&L)",
        "Interest",
        "Promoter Latest %", "Promoter Prev %", "Promoter Δ",
        "Error",
    ]
    cols   = [c for c in ordered if c in df.columns]
    extra  = [c for c in df.columns if c not in cols]
    df_out = df[cols + extra].copy()

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_out.to_excel(writer, index=False, sheet_name="Detail Analysis")
    buf.seek(0)

    wb = load_workbook(buf)
    ws = wb["Detail Analysis"]
    _hdr_style(ws)
    _row_style(ws)
    _autowidth(ws)
    _hyperlink_col(ws, "Link")

    # Map header → column index
    hdr_map: dict[str, int] = {
        str(cell.value): cell.column
        for cell in ws[1]
        if cell.value
    }

    # (col_name, low_threshold, high_threshold, reverse)
    color_rules = [
        ("ROE",             10,  20,  False),   # green ≥ 20, yellow ≥ 10
        ("ROCE",            10,  20,  False),
        ("P/BV",             1,   3,  True),    # lower better: green ≤ 1
        ("PAT Jumped CR",   10,  25,  False),
        ("OPM (Quarterly)", 10,  20,  False),
        ("OPM (P&L)",       10,  20,  False),
        ("Interest",        50, 100,  True),    # lower better: green ≤ 50
    ]

    for col_name, low, high, rev in color_rules:
        col_idx = hdr_map.get(col_name)
        if not col_idx:
            continue
        for ri in range(2, ws.max_row + 1):
            cell = ws.cell(row=ri, column=col_idx)
            v = parse_num(str(cell.value or ""))
            _color_cell(cell, v, low, high, reverse=rev)
            if v is not None:
                cell.number_format = "0.00"

    # Promoter Δ: green if buying (+), red if selling (-)
    col_idx = hdr_map.get("Promoter Δ")
    if col_idx:
        for ri in range(2, ws.max_row + 1):
            cell = ws.cell(row=ri, column=col_idx)
            v = parse_num(str(cell.value or ""))
            if v is not None:
                if v > 0:
                    cell.fill = PatternFill("solid", fgColor="C6EFCE")
                    cell.font = Font(color="276221", bold=True, name="Arial", size=10)
                elif v < 0:
                    cell.fill = PatternFill("solid", fgColor="FFC7CE")
                    cell.font = Font(color="9C0006", bold=True, name="Arial", size=10)
                cell.number_format = "+0.00;-0.00;0.00"

    ws.freeze_panes = "A2"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ── Core orchestrator ──────────────────────────────────────────────────────────
def run_full_scrape(
    session: requests.Session,
    max_pages: int,
    delay: float,
    on_page_start=None,
    on_page_done=None,
) -> tuple[pd.DataFrame, list]:
    """
    Scrape all results pages, filter for strictly-increasing Sales & Net Profit,
    and return (DataFrame, quarters).

    Optional callbacks for Streamlit progress:
      on_page_start(page, total, n_companies_so_far)
      on_page_done (page, total, n_companies_so_far)
    """
    first_html = fetch_page(session, 1)
    if not first_html:
        raise RuntimeError("Could not fetch page 1.")

    quarters = detect_quarters_from_html(first_html)
    if not quarters:
        raise RuntimeError("Could not detect quarters from HTML — check site structure.")

    total = get_total_pages(first_html)
    if max_pages and max_pages > 0:
        total = min(total, max_pages)

    all_companies: list = []
    if on_page_done:
        on_page_done(1, total, 0)
    all_companies.extend(parse_companies_from_html(first_html, quarters))

    for page in range(2, total + 1):
        if on_page_start:
            on_page_start(page, total, len(all_companies))
        time.sleep(delay * random.uniform(0.8, 1.5))
        html = fetch_page(session, page)
        all_companies.extend(parse_companies_from_html(html, quarters))
        if on_page_done:
            on_page_done(page, total, len(all_companies))

    qualifying: list = []
    for comp in all_companies:
        sales  = extract_metric(comp, "Sales",      quarters)
        profit = extract_metric(comp, "Net profit", quarters)
        if not is_strictly_increasing(sales,  quarters):
            continue
        if not is_strictly_increasing(profit, quarters):
            continue
        s = [sales[q]  for q in quarters]
        p = [profit[q] for q in quarters]
        if not s[0] or not p[0]:
            continue
        row = {"Company": comp["Company"], "Link": comp["Link"]}
        for i, q in enumerate(quarters):
            row[f"Sales {q}"]      = s[i]
            row[f"Net Profit {q}"] = p[i]
        row["Sales Growth %"]  = round((s[2] - s[0]) / abs(s[0]) * 100, 2)
        row["Profit Growth %"] = round((p[2] - p[0]) / abs(p[0]) * 100, 2)
        qualifying.append(row)

    return pd.DataFrame(qualifying), quarters